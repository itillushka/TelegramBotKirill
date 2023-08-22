import telebot
import openpyxl
from telebot import types
import webbrowser

TOKEN = '6633230318:AAEPmoWn2SgZsenyflbzZEP2hJ_Fgg6-diM'
broker_password = "1234"
bot = telebot.TeleBot(TOKEN)
DATA = "Data/user_data.xlsx"
CARGO = "Data/cargo_data.xlsx"

# Словарь для хранения данных о пользователях
user_data = {}
waiting_for_password = {}
chosen_cargo = {}
driver_data = {}

def is_user_registered(user_id):
    workbook = openpyxl.load_workbook(DATA)
    sheet = workbook.active
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] == user_id:
            role = row[1]
            workbook.close()
            return True, role
    workbook.close()
    return False, None


def get_user_data(user_id):
    workbook = openpyxl.load_workbook(DATA)
    sheet = workbook.active
    user_data = {}

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] == user_id:
            user_data = {
                "role": row[1],
                "name": row[2],
                "phone": row[3],
                "car_plate": row[4],
                "cargo_capacity": row[5],
                "dimensions": row[6],
                "body_type": row[7],
                "city": row[8],
                "distance": row[9],
                "ip_or_self_employed": row[10],
                "rent_or_own_car": row[11],
                "cargo_loading_type": row[12]  # Добавлен столбец для типа загрузки
            }
            break

    workbook.close()
    return user_data


# Обработчик команды /start
@bot.message_handler(commands=['start', 'menu'])
def start(message):
    markup = types.InlineKeyboardMarkup(row_width=1)
    broker_button = types.InlineKeyboardButton(" 🚚 Перевозчикам", callback_data="driver")
    driver_button = types.InlineKeyboardButton(" 📞 Диспетчерам", callback_data="broker")
    cargo_button = types.InlineKeyboardButton(" 📦 Отправить груз", callback_data="cargo")
    markup.add(broker_button, driver_button, cargo_button)
    bot.send_message(message.chat.id, "Приветствую в нашем боте! Пожалуйста выберите раздел:", reply_markup=markup)


@bot.callback_query_handler(func=lambda call: call.data == "broker")
def handle_broker_role(call):
    user_id = call.from_user.id
    bot.send_message(user_id, "Спасибо, что хотите пополнить нашу команду диспетчеров!\nПрошу вас заполнить Гугл форму, чтобы мы узнали о вас побольше!")

    # Создаем кнопку для перенаправления на сайт Google
    markup = types.InlineKeyboardMarkup(row_width=1)
    google_button = types.InlineKeyboardButton("Перейти к Гугл форме", url="https://www.google.com")
    markup.add(google_button)

    # Отправляем сообщение с кнопкой
    bot.send_message(user_id, "Для заполнения формы, перейдите по ссылке ниже:", reply_markup=markup)


@bot.callback_query_handler(func=lambda call: call.data == "driver")
def handle_driver_role(call):
    user_id = call.from_user.id
    registered, user_role = is_user_registered(user_id)

    if registered and user_role == "Водитель":
        markup = types.InlineKeyboardMarkup(row_width=1)
        my_data_button = types.InlineKeyboardButton("Мои данные", callback_data="my_data")
        view_cargo_button = types.InlineKeyboardButton("Посмотреть грузы", callback_data="view_cargo")
        markup.add(my_data_button, view_cargo_button)
        bot.send_message(user_id, "Добро пожаловать в меню водителя", reply_markup=markup)
    elif registered and user_role == "Брокер":
        bot.send_message(user_id, "Вы не имеете доступа к роли Перевозчика.")
    elif not registered:
        start_button = types.InlineKeyboardButton("[🟢 Начать ]", callback_data="start_driver")
        markup = types.InlineKeyboardMarkup(row_width=1)
        markup.add(start_button)
        bot.send_message(user_id, "Добро пожаловать в панель новых водителей!", reply_markup=markup)


@bot.callback_query_handler(func=lambda call: call.data in ["start_driver"])
def start_driver(call):
    user_id = call.from_user.id
    status = call.data
    if status == "start_driver":
        user_data[user_id] = {"role": "Водитель"}
        driver_data[user_id] = {}  # Создаем пустой словарь для данных водителя
        bot.send_message(user_id, "Отлично! Пожалуйста, ответьте на несколько вопросов.")
        bot.send_message(user_id, "Ваше полное имя?")
        bot.register_next_step_handler(call.message, ask_phone)


@bot.callback_query_handler(func=lambda call: call.data in ["my_data", "view_cargo"])
def handle_driver_choice(call):
    user_id = call.from_user.id
    choice = call.data

    if choice == "my_data":
        user_data = get_user_data(user_id)

        if user_data:
            response = "👤 Ваши данные:\n"
            for key, value in user_data.items():
                response += f"✅ {key.capitalize()}: {value}\n"
            bot.send_message(user_id, response)
        else:
            bot.send_message(user_id, "🚫 Ваши данные не найдены.")
    elif choice == "view_cargo":
        cargo_workbook = openpyxl.load_workbook(CARGO)
        cargo_sheet = cargo_workbook.active

        cargo_buttons = []
        for row in cargo_sheet.iter_rows(min_row=2, values_only=True):
            from_location = row[0]
            to_location = row[1]

            cargo_buttons.append(types.InlineKeyboardButton(f"Груз: {from_location} -> {to_location}",
                                                            callback_data=f"cargo_{from_location}_{to_location}"))

        # Добавляем кнопку "Готово" в конце списка грузов
        finish_button = types.InlineKeyboardButton("Готово✅", callback_data="finish")
        cargo_buttons.append(finish_button)

        cargo_buttons_markup = types.InlineKeyboardMarkup(row_width=1)
        cargo_buttons_markup.add(*cargo_buttons)

        bot.send_message(user_id, "Выберите груз:", reply_markup=cargo_buttons_markup)


@bot.callback_query_handler(func=lambda call: call.data.startswith("cargo_"))
def handle_cargo_choice(call):
    user_id = call.from_user.id
    cargo_data = call.data.split("_")[1:]  # Разделяем данные о грузе из callback_data
    cargo_key = "_".join(cargo_data)  # Создаем уникальный ключ для груза

    if user_id not in chosen_cargo:
        chosen_cargo[user_id] = []

    if cargo_key == "finish":
        # Если выбрана кнопка "Готово", обрабатываем завершение выбора грузов
        if user_id in chosen_cargo and chosen_cargo[user_id]:
            chosen_cargo_rows = chosen_cargo[user_id]

            # Добавление номеров строк выбранных грузов в столбец "Груз и номер груза"
            for cargo_row in chosen_cargo_rows:
                add_chosen_cargo(user_id, cargo_row)

            chosen_cargo[user_id] = []  # Очищаем список выбранных грузов
            bot.send_message(user_id, "Спасибо за выбор! Мы с вами свяжемся. 🚚")
        else:
            bot.send_message(user_id, "Вы еще не выбрали грузы. 🚫")
    else:
        # Иначе, добавляем выбранный груз
        chosen_cargo[user_id].append(cargo_key)
        bot.answer_callback_query(call.id, text="Груз добавлен в выбранные! ✅")


@bot.callback_query_handler(func=lambda call: call.data == "finish")
def handle_finish(call):
    user_id = call.from_user.id
    if user_id in chosen_cargo and chosen_cargo[user_id]:
        chosen_cargo_rows = chosen_cargo[user_id]

        # Добавление номеров строк выбранных грузов в столбец "Груз и номер груза"
        for cargo_row in chosen_cargo_rows:
            add_chosen_cargo(user_id, cargo_row)

        chosen_cargo[user_id] = []  # Очищаем список выбранных грузов
        bot.send_message(user_id, "Спасибо за выбор! Мы с вами свяжемся. 🚚")
    else:
        bot.send_message(user_id, "Вы еще не выбрали грузы. 🚫")


def add_chosen_cargo(user_id, cargo_row):
    workbook = openpyxl.load_workbook(DATA)
    sheet = workbook.active

    # Найдем строку, где user_id совпадает
    for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if row[0] == user_id:
            user_row = idx  # Индекс строки, где нашли совпадение

            # Получим номер последнего заполненного столбца в текущей строке
            last_column = sheet.max_column

            # Запишем груз в следующий столбец
            sheet.cell(row=user_row, column=last_column + 1, value=cargo_row)
            break  # Прерываем цикл, так как нашли нужную строку

    workbook.save(DATA)
    workbook.close()


def ask_phone(message):
    user_id = message.from_user.id
    driver_data[user_id]["name"] = message.text

    bot.send_message(message.chat.id, "Ваш телефон?")
    bot.register_next_step_handler(message, ask_car_number)


def ask_car_number(message):
    user_id = message.from_user.id
    driver_data[user_id]["phone"] = message.text

    bot.send_message(message.chat.id, "Государственный знак машины?")
    bot.register_next_step_handler(message, ask_cargo_capacity)


def ask_cargo_capacity(message):
    user_id = message.from_user.id
    driver_data[user_id]["car_number"] = message.text

    bot.send_message(message.chat.id, "Грузоподъемность машины?")
    bot.register_next_step_handler(message, ask_dimensions)


def ask_dimensions(message):
    user_id = message.from_user.id
    driver_data[user_id]["cargo_capacity"] = message.text

    bot.send_message(message.chat.id, "Длина/Ширина/Высота машины?")
    bot.register_next_step_handler(message, ask_body_type)


def ask_body_type(message):
    user_id = message.from_user.id
    driver_data[user_id]["dimensions"] = message.text

    bot.send_message(message.chat.id, "Тип кузова?")
    bot.register_next_step_handler(message, ask_residence_city)


def ask_residence_city(message):
    user_id = message.from_user.id
    driver_data[user_id]["body_type"] = message.text

    bot.send_message(message.chat.id, "Город проживания?")
    bot.register_next_step_handler(message, ask_distance_to_travel)


def ask_distance_to_travel(message):
    user_id = message.from_user.id
    driver_data[user_id]["residence_city"] = message.text

    bot.send_message(message.chat.id, "Дистанция, на которую готовы ездить? (в километрах)")
    bot.register_next_step_handler(message, ask_employment_type)


def ask_employment_type(message):
    user_id = message.from_user.id
    driver_data[user_id]["distance_to_travel"] = message.text

    bot.send_message(message.chat.id, "Вы являетесь ИП или самозанятым?")
    bot.register_next_step_handler(message, ask_car_ownership)


def ask_car_ownership(message):
    user_id = message.from_user.id
    driver_data[user_id]["employment_type"] = message.text

    bot.send_message(message.chat.id, "У вас машина в аренде или личная?")
    bot.register_next_step_handler(message, save_driver_info)


def save_driver_info(message):
    user_id = message.from_user.id
    driver_data[user_id]["car_ownership"] = message.text

    # Добавление информации о водителе в Excel таблицу
    add_driver_to_excel(user_id, **driver_data[user_id])

    bot.send_message(message.chat.id, "Какой у вас тип загрузки? Задний/верхний/боковой?")
    bot.register_next_step_handler(message, ask_cargo_loading_type)

def ask_cargo_loading_type(message):
    user_id = message.from_user.id
    driver_data[user_id]["cargo_loading_type"] = message.text

    add_driver_to_excel(user_id, **driver_data[user_id])  # Добавляем данные о водителе в таблицу
    bot.send_message(user_id, "Спасибо! Ваши данные сохранены.")


def add_driver_to_excel(user_id, **data):
    workbook = openpyxl.load_workbook(DATA)
    sheet = workbook.active

    # Извлекаем данные о водителе из словаря и добавляем в строку
    driver_info = [
        user_id,
        "Водитель",
        data["name"],
        data["phone"],
        data["car_number"],
        data["cargo_capacity"],
        data["dimensions"],
        data["body_type"],
        data["residence_city"],
        data["distance_to_travel"],
        data["employment_type"],
        data["car_ownership"],
        data["cargo_loading_type"]  # Добавлен столбец для типа загрузки
    ]
    sheet.append(driver_info)

    workbook.save(DATA)
    workbook.close()




@bot.callback_query_handler(func=lambda call: call.data == "cargo")
def handle_cargo(call):
    user_id = call.from_user.id
    user_data[user_id] = {}
    bot.send_message(user_id, "Введите данные о грузе.\n\n1. Откуда?")

    bot.register_next_step_handler(call.message, ask_cargo_from)


def ask_cargo_from(message):
    user_id = message.from_user.id
    user_data[user_id]["cargo_from"] = message.text
    bot.send_message(user_id, "2. Куда?")
    bot.register_next_step_handler(message, ask_cargo_to)


def ask_cargo_to(message):
    user_id = message.from_user.id
    user_data[user_id]["cargo_to"] = message.text
    bot.send_message(user_id, "3. Дистанция (в километрах)?")
    bot.register_next_step_handler(message, ask_cargo_distance)


def ask_cargo_distance(message):
    user_id = message.from_user.id
    user_data[user_id]["cargo_distance"] = message.text
    bot.send_message(user_id, "4. Вес груза (в кг)?")
    bot.register_next_step_handler(message, ask_cargo_weight)


def ask_cargo_weight(message):
    user_id = message.from_user.id
    user_data[user_id]["cargo_weight"] = message.text
    bot.send_message(user_id, "5. Оплата (в рублях)?")
    bot.register_next_step_handler(message, save_cargo_info)


def save_cargo_info(message):
    user_id = message.from_user.id
    cargo_info = {
        "from_location": user_data[user_id]["cargo_from"],
        "to_location": user_data[user_id]["cargo_to"],
        "distance": user_data[user_id]["cargo_distance"],
        "weight": user_data[user_id]["cargo_weight"],
        "payment": message.text
    }

    add_cargo_to_excel(**cargo_info)
    bot.send_message(user_id, "Спасибо! Данные о грузе сохранены.")


def add_cargo_to_excel(from_location, to_location, distance, weight, payment):
    workbook = openpyxl.load_workbook(CARGO)
    sheet = workbook.active

    sheet.append([from_location, to_location, distance, weight, payment])
    workbook.save(CARGO)
    workbook.close()


bot.polling()
